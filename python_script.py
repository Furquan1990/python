import xlrd
import pandas as pd
import openpyxl
import datetime
import os
import shutil
date = datetime.date.today()
newpath = 'C:/MOS Script/'+str(date)
if os.path.exists(newpath):
    shutil.rmtree(newpath)
    os.makedirs(newpath)
    os.startfile('C:/MOS Script/'+str(date)) 
else:
    os.makedirs(newpath)
    os.startfile('C:/MOS Script/'+str(date)) 
xl = pd.ExcelFile('C:/MOS Script/input.xlsx')
df = xl.parse('Sheet1')
df1 = df.sort_values(by=['Site'],ascending=True)
excel_handler = pd.ExcelWriter('C:/MOS Script/data2.xlsx')
df1.to_excel(excel_handler,sheet_name='Sheet1')
excel_handler.save()
filename = 'C:/MOS Script/data2.xlsx'
xlsx_file = openpyxl.load_workbook(filename)
sheet = xlsx_file['Sheet1']
row_count = sheet.max_row
column_count = sheet.max_column
i = 2
j = 3
for i in range(2,row_count+1):
    if sheet.cell(i,2).value == sheet.cell(i+1,2).value:
        site = sheet.cell(i,2).value
        with open('C:/MOS Script/'+str(date)+'/'+ str(site)+'.txt', 'a+') as f:
            j = 3
            for j in range(3,column_count+1):
                print(sheet.cell(i,j).value, file = f)           
    else:
        site = sheet.cell(i,2).value
        with open('C:/MOS Script/'+str(date)+'/'+ str(site)+'.txt', 'a+') as f:
            j = 3
            for j in range(3,column_count+1):
                print(sheet.cell(i,j).value, file = f)
os.remove(filename)
